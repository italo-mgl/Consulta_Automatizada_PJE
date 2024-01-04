from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

numero_oab = 133864

#Entrar no site da - "https://pje-consulta-publica.tjmg.jus.br"
driver = webdriver.Chrome()
driver.get("https://pje-consulta-publica.tjmg.jus.br")
sleep(5)

#Digitar numero OAB 
campo_oab = driver.find_element(By.XPATH,"//*[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)

#selecionar estado
estado = driver.find_element(By.XPATH,'//*[@id="fPP:Decoration:estadoComboOAB"]')
opcoes_estados = Select(estado)
opcoes_estados.select_by_visible_text("SP")

#Clicar em pesquisar
pesquisar = driver.find_element(By.XPATH,'//*[@id="fPP:searchProcessos"]')
pesquisar.click()
sleep(10)

#Entrar em cada um dos processos
processos = driver.find_elements(By.XPATH,"//b[@class='btn-block']")
for processos in processos:
    processos.click()
    sleep(10)
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    driver.set_window_size(1280,720)

    #Extrair n° do processo
    numero_processo = driver.find_elements(By.XPATH,"//div[@class='col-sm-12 ']")
    numero_processo = numero_processo[0]
    numero_processo = numero_processo.text

    #Extrair data
    data_distribuicao = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")
    data_distribuicao = data_distribuicao[1]
    data_distribuicao = data_distribuicao.text
    sleep(5)

    #Extrair e guardar todas as ultimas movimentações
    movimentacoes = driver.find_elements(By.XPATH,'//div[@id="j_id132:processoEventoPanel_body"]//tr[contains(@class,"rich-table-row")]//td//div//div//span')
    lista_movimentacoes = []
    for movimentacao in movimentacoes:
        lista_movimentacoes.append(movimentacao.text)

#Guardar tudo no excel, separados por processo
    workbook = openpyxl.load_workbook("dados.xlsx")
    try:
        #Codigo para inserir em pagina existente
        workbook.create_sheet(numero_processo)

        #Acessar pagina processo
        pagina_processo = workbook[numero_processo]

        #Criar nome colunas
        pagina_processo['A1'].value = "Numero Processo"
        pagina_processo['B1'].value = "Data Distribuição"
        pagina_processo['C1'].value = "Movimentações"

        #Adicionar numero do processo
        pagina_processo["A2"].value = numero_processo
        
        #Adicionar data de distribuição
        pagina_processo["B2"].value = data_distribuicao

        #Adicionar movimentações
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2,max_row=len(lista_movimentacoes),min_col=3,max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save("dados.xlsx")
        driver.close()
        sleep(5)
        driver.switch_to.window(driver.window_handles[0])

        
    except Exception as error:
        #Codigo para criar uma página do zero e inserir informações
        workbook.create_sheet(numero_processo)

        #Acessar pagina processo
        pagina_processo = workbook[numero_processo]

        #Criar nome colunas
        pagina_processo['A1'].value = "Numero Processo"
        pagina_processo['B1'].value = "Data Distribuição"
        pagina_processo['C1'].value = "Movimentações"

        #Adicionar numero do processo
        pagina_processo["A2"].value = numero_processo
        
        #Adicionar data de distribuição
        pagina_processo["B2"].value = data_distribuicao

        #Adicionar movimentações
        pagina_processo["C2"].value = movimentacoes
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2,max_row=len(lista_movimentacoes),min_col=3,max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save("dados.xlsx")
        driver.close()
        sleep(5)
        driver.switch_to.window(driver.window_handles[0])
        
